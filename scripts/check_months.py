import openpyxl
wb = openpyxl.load_workbook('/Users/brunobarros/Projects/axpo/sim_open.xlsm', read_only=True, data_only=True)
ws = wb['.']
print('dot sheet rows 24-36 cols D and E:')
for i, row in enumerate(ws.iter_rows(min_row=24, max_row=36, values_only=True), 24):
    print(i, repr(row[3]), '|', repr(row[4]))  # D=index3, E=index4

print('\nDINAMICA N1 rows 44-57 col A and AQ(43) and AT(46):')
ws2 = wb['DINAMICA N1']
for i, row in enumerate(ws2.iter_rows(min_row=44, max_row=57, values_only=True), 44):
    a = row[0]
    aq = row[42] if len(row) > 42 else None
    at = row[45] if len(row) > 45 else None
    print(i, repr(a), '|AQ=', repr(aq), '|AT=', repr(at))
