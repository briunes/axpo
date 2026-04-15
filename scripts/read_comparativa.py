import openpyxl

# Check BASE DE DATOS FIJO in sim_open.xlsm vs open_file.xlsm
print("=== sim_open.xlsm - BASE DE DATOS FIJO (2.0TD prices) ===")
wb = openpyxl.load_workbook("/Users/brunobarros/Projects/axpo/sim_open.xlsm", read_only=True, data_only=True)
ws = wb["BASE DE DATOS FIJO"]
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
    vals = [v for v in row if v is not None]
    if vals and any(str(v) in ['2.0TD', 'ESTABLE N1', 'POTENCIA', 'P1', 'P2', 'P3'] for v in vals[:3]):
        all_vals = list(row)
        non_none = [(idx, v) for idx, v in enumerate(all_vals) if v is not None]
        print(f"Row {i}: {non_none[:12]}")

print("\n=== sim_open.xlsm - '.' sheet row 69 (2.0TD FEBRERO-25 ESTABLE N1) ===")
ws2 = wb["."]
# Find FEBRERO-25 ESTABLE N1 in the lookup table
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
    vals = list(row)
    first = vals[0] if vals else None
    if first and "FEBRERO" in str(first) and "2.0TD" in str(first) and ("ESTABLE" in str(first) or "DINAMICA" in str(first)):
        non_none = [(idx, v) for idx, v in enumerate(vals) if v is not None]
        print(f"Row {i}: {non_none}")

print("\n=== sim_open.xlsm - '.' sheet COMPARADOR section ===")
for i, row in enumerate(ws2.iter_rows(min_row=1260, max_row=1340, values_only=True), start=1260):
    vals = [v for v in row if v is not None]
    if vals:
        print(f"Row {i}: {vals[:10]}")

print("\n=== open_file.xlsm - BASE DE DATOS FIJO (2.0TD prices) ===")
wb2 = openpyxl.load_workbook("/Users/brunobarros/Projects/axpo/open_file.xlsm", read_only=True, data_only=True)
ws3 = wb2["BASE DE DATOS FIJO"]
for i, row in enumerate(ws3.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
    vals = [v for v in row if v is not None]
    if vals and any(str(v) in ['2.0TD', 'ESTABLE N1', 'POTENCIA', 'P1', 'P2', 'P3'] for v in vals[:3]):
        all_vals = list(row)
        non_none = [(idx, v) for idx, v in enumerate(all_vals) if v is not None]
        print(f"Row {i}: {non_none[:12]}")

# Read row 6 (2.0TD ENERO-26 ESTABLE N1) for comparison
print("\n=== FULL CONTENT OF ROW 6 ===")
for i, row in enumerate(ws.iter_rows(min_row=6, max_row=6, values_only=True), start=6):
    all_vals = list(row)
    non_none = [(idx, v) for idx, v in enumerate(all_vals) if v is not None]
    print(f"Non-None values with indices: {non_none}")

# Header row
print("\n=== HEADER ROW 1 ===")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
    all_vals = list(row)
    non_none = [(idx, v) for idx, v in enumerate(all_vals) if v is not None]
    if non_none:
        print(f"Row {i}: {non_none[:20]}")

# Also read the COMPARATIVA LUZ sheet
print("\n=== COMPARATIVA LUZ SHEET ===")
ws2 = wb["COMPARATIVA LUZ"]
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
    vals = [v for v in row if v is not None]
    if vals:
        all_vals = list(row)
        non_none = [(idx, v) for idx, v in enumerate(all_vals) if v is not None]
        print(f"Row {i}: {non_none[:15]}")
