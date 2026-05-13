import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Incidencias"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER_WRAP = Alignment(wrap_text=True, vertical="top", horizontal="center")

FILL_HEADER   = PatternFill("solid", fgColor="1F3864")
FILL_RESOLVED = PatternFill("solid", fgColor="C6EFCE")
FILL_OPEN     = PatternFill("solid", fgColor="FFEB9C")
FILL_IGNORED  = PatternFill("solid", fgColor="E0E0E0")

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
FONT_RESOLVED = Font(color="276221", bold=True, size=10)
FONT_OPEN     = Font(color="9C6500", bold=True, size=10)
FONT_IGNORED  = Font(color="595959", bold=True, size=10)
FONT_NORMAL   = Font(size=10)

STATUS_LABEL = {
    "closed":  "Resuelto",
    "open":    "En curso",
    "ignored": "No procede",
}

headers = ["N", "Incidencia", "Descripcion", "Comentarios del cliente", "Reportado por", "Estado", "Respuesta Axpo"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[1].height = 28

rows = [
  (1,  "Acceso Jesus Millan", "No me deja acceder con el nombre y usuario que me ha llegado al email.", "", "Jesus Millan", "ignored", "Las credenciales fueron enviadas correctamente. El acceso fue verificado y funciona. Se descarta la incidencia."),
  (2,  "Acceso Eva Morales", "No me deja acceder con el nombre y usuario que me ha llegado al email.", "", "Eva Morales", "ignored", "Las credenciales fueron enviadas correctamente. El acceso fue verificado y funciona. Se descarta la incidencia."),
  (3,  "Importe Ahorro", "No coincide el importe del ahorro con la tarifa seleccionada.", "", "Ruth Chucad", "closed", "Corregido. El importe de ahorro ahora coincide correctamente con la tarifa seleccionada."),
  (4,  "Conceptos de Factura", "No son correctas los terminos al traducir: Potencia Contraida, Potencia Excedente, Alquiler Parquimetro, Historial de Descarga.", "", "Ruth Chucad", "closed", "Corregido. Los terminos de los conceptos de factura han sido actualizados: Potencia Contraida, Potencia Excedente, Alquiler de Equipo e Historial de Precios."),
  (5,  "Precios Potencia", "No tienen coherencia de donde proceden los precios de potencia de la competencia. Se esta indicando un Coste Anual en Potencia con mucha diferencia.", "", "Ruth Chucad", "open", "En revision. Se esta analizando la coherencia de los precios de potencia de la competencia y la diferencia en el Coste Anual de Potencia."),
  (6,  "Consumo anual y tarifa", "OCR parece que no calcula el consumo anual de la factura. No rellena la tarifa de acceso.", "Son datos que se pueden obtener y acabar de completar a mano.", "Jose Campo", "closed", "El consumo anual y la tarifa de acceso pueden completarse manualmente cuando el OCR no los extrae. Resuelto."),
  (7,  "Parrilla simulaciones", "Buena parrilla de presentacion de las simulaciones realizadas.", "Se parece mucho a una pantalla de gestion de correo electronico.", "Jose Campo", "ignored", "Diseno revisado. La pantalla de simulaciones tiene su propia identidad visual y no requiere cambios adicionales."),
  (8,  "Comisiones por operacion", "Seria interesante que incluyera las comisiones por la operacion, pudiendo ser configurable segun el perfil.", "Riesgo: con un mismo perfil lo usen personas que no deben conocer el dato de comision.", "Jose Campo", "ignored", "Funcionalidad de comisiones aplazada. Riesgo identificado: usuarios con el mismo perfil podrian ver datos de comision ajenos, responsabilidad del agente."),
  (9,  "Clientes", "Buen acceso a la lista de clientes. Que clientes se incluyen aqui? Es posible descargar aqui los leads que le proporcionamos a los agentes?", "", "Jose Campo", "ignored", "Cada agencia ve unicamente sus propios clientes. El administrador puede ver clientes de todas las agencias. Los comerciales ven solo los suyos. La importacion de leads puede valorarse en una fase futura."),
  (10, "Sugerencia precio kWh", "Reflejar los precios en Termino de Energia en euros/kWh y no solo el Coste Total.", "", "Ruth Chucad", "ignored", "Sugerencia registrada. Los precios en euros/kWh no se mostraran en esta version. Puede considerarse en futuras iteraciones."),
  (11, "Ordenar por ahorro", "Una vez calculada la comparativa, ordenar las tarifas en funcion del porcentaje de ahorro.", "", "Ruth Chucad", "closed", "Resuelto. Los resultados de la comparativa se ordenan ahora de mayor a menor porcentaje de ahorro."),
  (12, "Comisiones en resultados", "Deben aparecer las comisiones de cada producto o tarifa simulado.", "Los sub agentes se pueden identificar con el % que desee el master, igual que en Salesforce en la comision estimada.", "Jorge Lopez", "ignored", "La visualizacion de comisiones en los resultados queda aplazada para la siguiente fase del proyecto."),
  (13, "Datos Cliente", "Cuando editas a mano no marca en todos los campos del nombre del cliente dicho nombre, debes editar cada uno.", "", "Jorge Lopez", "closed", "Corregido. Al editar el nombre del cliente, todos los campos relacionados se actualizan automaticamente."),
  (14, "Borrado de fechas", "No puedes borrar todo el campo de fecha de una vez.", "", "Jorge Lopez", "closed", "Corregido. Los campos de fecha pueden borrarse completamente de una sola vez."),
  (15, "Importe Factura", "Al marcar la coma en el importe, te edita de la siguiente manera 54,00045, y deberia ser 54,45.", "", "Jorge Lopez", "closed", "Corregido. El campo de importe gestiona correctamente el separador decimal con coma (ej. 54,45)."),
  (16, "Nombre de Agente", "El nombre de agente debe aparecer por usuario y no tener que editar manualmente.", "", "Jorge Lopez", "closed", "Corregido. El nombre del agente se autocompleta con el usuario que ha iniciado sesion."),
  (17, "PDF simulacion", "El PDF no aparecia o daba fallo en descarga.", "", "Jorge Lopez", "closed", "Resuelto. La generacion y descarga del PDF funciona correctamente."),
  (18, "Productos Gas", "No aparecen los productos Gas Estable N2 y N3. Ademas estan denominados como Gas Fijo, nombre que no existe como producto de campana.", "", "Jorge Lopez", "closed", "Resuelto. Se han anadido los productos Gas Estable N2 y Gas Estable N3. La denominacion Gas Fijo ha sido sustituida por los nombres de campana correctos."),
  (19, "Lectura OCR - campos", "Los datos los posiciona en celdas que no corresponden.", "", "Jorge Lopez", "closed", "Resuelto. El OCR posiciona ahora los datos extraidos en los campos correctos."),
  (20, "OCR - Drag and Drop", "Mi primer impulso fue arrastrar la factura al hueco de subida, pero no esta implementado. Si no es muy costoso, se podria poner.", "", "Francisco Zarzuela", "closed", "Resuelto. Se ha implementado la subida de facturas mediante arrastrar y soltar (drag and drop)."),
  (21, "OCR - CUPS Naturgy", "Naturgy electricidad no me ha codificado el CUPS.", "", "Francisco Zarzuela", "closed", "Resuelto. El CUPS se extrae correctamente de las facturas de Naturgy electricidad."),
  (22, "Estructura datos factura", "Seccion 3 Datos de Factura deja algunos campos sueltos mas abajo. Lo englobaria todo en un cajon.", "", "Francisco Zarzuela", "closed", "Resuelto. Todos los campos relacionados estan ahora agrupados en una unica seccion colapsable Desglose de Factura."),
  (23, "Perfil de Carga duplicado", "Perfil de Carga aparece dos veces. No se cual es el vinculado. Lo quitaria de Datos de Factura y explicaria bien que implica marcarlo.", "", "Francisco Zarzuela", "closed", "Resuelto. El Perfil de Carga aparece ahora una unica vez, dentro de la subseccion de Datos de Factura, con descripcion de su funcion."),
  (24, "Opciones personalizadas", "Las opciones personalizadas no deberian estar ahi; si necesito una personalizada tengo que ir atras. Las pondria en la seccion 4 junto con la tabla de resultados.", "", "Francisco Zarzuela", "closed", "Resuelto. Las opciones personalizadas se muestran directamente en la pestana de Resultados, asociadas a cada fila correspondiente."),
  (25, "Fijo personalizado", "Hay que buscar un hueco para un fijo personalizado, que sea libre y directamente meta precios de energia y potencia.", "", "Francisco Zarzuela", "closed", "Resuelto. Se ha anadido la opcion Fijo Personalizado que permite introducir precios de energia y potencia libremente."),
  (26, "Seccion Resultados", "Al consumo no le veo mucho sentido de tocar ahi. OMIE euros/MWh no se lo que hace. Aqui pondria las opciones de personalizada, ligado a una linea de resultados.", "", "Francisco Zarzuela", "closed", "Resuelto. El campo OMIE euros/MWh ha sido retirado. Las opciones personalizadas estan ahora en la seccion de Resultados."),
  (27, "Historico precios", "Se repiten en todos los meses al seleccionar tarifa dinamica control.", "", "Francisco Zarzuela", "closed", "Resuelto. El historico de precios muestra correctamente los valores mensuales sin repeticiones al seleccionar Dinamica Control."),
  (28, "Descarga PDF", "Deberia estar mas claro donde se descarga la simulacion. El boton Compartir deja dudas.", "", "Francisco Zarzuela", "closed", "Resuelto. La ubicacion del boton de descarga del PDF esta mas clara. El flujo de compartir/descargar ha sido simplificado."),
  (29, "Plantillas personalizadas", "Dar opcion a personalizar las plantillas, incluso subir logo de agencia para que aparezca en la plantilla.", "", "Francisco Zarzuela", "ignored", "Las plantillas PDF personalizadas por agencia, incluyendo subida de logo, se valoraran en una fase futura."),
  (30, "Diseno simulacion final", "No tiene nada que ver con el diseno del Excel. Aparece en dos columnas practicamente espejo.", "", "Francisco Zarzuela", "ignored", "El diseno de la simulacion final es intencional y adaptado al entorno web. No esta basado en el Excel de referencia."),
  (31, "Vincular consumos Nemon", "Si hubiera posibilidad de reflejar el consumo anual automaticamente en cuanto se introduce el CUPS.", "", "Ruth Chucad", "ignored", "La integracion con Nemon para obtener el consumo anual al introducir el CUPS queda pendiente para una fase futura."),
  (32, "OCR - CUPS Iberdrola", "Quiza sea porque en las facturas de Iberdrola vienen separados los digitos, pero se tiene que introducir cada digito del CUPS.", "", "Ruth Chucad", "closed", "Resuelto. El CUPS se autocompleta correctamente, incluso en facturas de Iberdrola con digitos separados."),
  (33, "OCR - Excesos Potencia", "Aunque no sea campo obligatorio, deberia completarse si viene claramente reflejado en factura.", "", "Ruth Chucad", "closed", "Resuelto. Los excesos de potencia se rellenan automaticamente desde la factura cuando estan claramente reflejados."),
  (34, "PDF - comparativa", "Cuando se compara Plano Actual con Plano Axpo, las dos primeras filas comparan los mismos datos. Deberia ser mas relevante indicar los precios.", "", "Ruth Chucad", "ignored", "Las dos primeras filas de la comparativa PDF muestran los mismos datos intencionadamente. Se valorara redinsenar esta seccion en una version futura."),
  (35, "Potencia - campo edicion", "Al registrar la potencia sale por defecto el 0, no da opcion a borrarlo y escribir la potencia correcta.", "Que exista la opcion de escribir la potencia tal cual.", "Ana", "closed", "Resuelto. El campo de potencia permite borrar el valor por defecto (0) y escribir directamente la potencia correcta."),
  (36, "IVA e Impuesto Electrico", "Pienso que es mejor tener una check list con todas las opciones.", "", "Ana", "closed", "Resuelto. El IVA y el Impuesto Electrico se presentan como selectores desplegables con opciones configurables desde los Ajustes del Sistema."),
  (37, "Load Profile duplicado", "Esta duplicado. Lo pide dos veces.", "", "Ana", "closed", "Resuelto. El Perfil de Carga duplicado ha sido eliminado. Ahora aparece unicamente en la seccion de Datos de Factura."),
  (38, "Insertar precio 1PT", "Producto para TLV.", "", "Ana", "closed", "Resuelto. Los productos 1P_PLUS y 1P_PLUS_XL son de Periodo Unico. El precio P1 se aplica a todos los periodos, funcionando correctamente con la tarifa 6.1TD (TLV)."),
  (39, "Pestana Simulaciones", "La ultima columna no se que es lo que pone. No puede haber una opcion en esta misma vista que descargue la simulacion en PDF?", "", "Ana", "closed", "Resuelto. La ultima columna de la lista de simulaciones ha sido clarificada. Se ha anadido una opcion de descarga directa del PDF desde esa vista."),
  (40, "Re-entrar en simulacion", "He hecho la simulacion, he escogido un producto y en simulaciones no puedo entrar en la misma. No lo veo claro.", "", "Ana", "ignored", "El comportamiento es el previsto: una vez que la simulacion ha sido compartida con el cliente, no puede modificarse para garantizar la integridad de la oferta enviada."),
  (41, "Anadir aviso en PDF", "Que no se garantizan estos precios.", "", "Ana", "closed", "Resuelto. El PDF incluye ahora el aviso: Los precios indicados no estan garantizados."),
  (42, "Grafico visual ahorro", "Anadir graficos visuales en el ahorro (anual y en ciclo facturacion).", "", "Ana", "closed", "Resuelto. Se han anadido graficos visuales del ahorro anual y por ciclo de facturacion en la vista de resultados."),
  (43, "Lectura OCR - error", "Con mi usuario no me lee ninguna factura, me da un error de intentelo mas tarde pero nunca las lee. He probado varios formatos.", "", "David Garcia", "closed", "Resuelto. El error de lectura OCR ha sido corregido. El sistema lee correctamente las facturas en multiples formatos."),
  (44, "PDF - ahorro anual", "En el resultado final de la simulacion al exportar PDF, noto a faltar el porcentaje anual de ahorro y monetizar ese ahorro anual.", "", "David Garcia", "closed", "Resuelto. El PDF exportado incluye ahora el porcentaje de ahorro anual y el importe monetizado del ahorro anual."),
  (45, "Visibilidad simulaciones", "En los accesos que demos a las agencias, entiendo que solo visualizaran sus simulaciones y no todas las generadas.", "", "David Garcia", "ignored", "Confirmado. Los usuarios de agencia ven unicamente sus propias simulaciones. Los agentes ven las de su agencia. Los comerciales ven solo las que ellos han creado."),
  (46, "Importe potencia incorrecto", "Mal calculado importe total potencia (Panificadora Amaya).", "", "", "open", "En revision. Se esta analizando el calculo incorrecto del importe total de potencia para el caso Panificadora Amaya."),
  (47, "Mes de facturacion", "No detecta bien el mes de facturacion.", "", "Sergio", "closed", "Resuelto. El simulador detecta correctamente el mes de facturacion desde la factura."),
  (48, "Precios de campana", "Ojo a la actualizacion de precios de campana.", "", "Sergio", "ignored", "Aviso registrado. Los precios de campana se monitorizan y actualizan periodicamente. Es responsabilidad del equipo mantenerlos al dia."),
  (49, "Historico de Precios", "Historico de Precios (revisar datos).", "", "Sergio", "closed", "Resuelto. Los datos del historico de precios han sido revisados y corregidos."),
  (50, "Periodos de facturacion", "No se pueden modificar manualmente los periodos de facturacion.", "No se pueden modificar manualmente los periodos de facturacion.", "Carmen", "closed", "Resuelto. Los periodos de facturacion pueden editarse manualmente."),
  (51, "OCR - multiples archivos", "Solo sube un JPEG, deberia subir multiples.", "Solo sube un JPEG, deberia subir multiples.", "Samuel", "closed", "Resuelto. La subida de facturas admite ahora multiples archivos en distintos formatos (no solo un JPEG)."),
]

for i, (num, incidencia, desc, comentario, reporter, status, respuesta) in enumerate(rows, 2):
    data = [num, incidencia, desc, comentario, reporter, STATUS_LABEL[status], respuesta]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=i, column=col, value=val)
        c.font = FONT_NORMAL
        c.alignment = WRAP
        c.border = BORDER

    estado = ws.cell(row=i, column=6)
    if status == "closed":
        estado.fill = FILL_RESOLVED
        estado.font = FONT_RESOLVED
    elif status == "open":
        estado.fill = FILL_OPEN
        estado.font = FONT_OPEN
    else:
        estado.fill = FILL_IGNORED
        estado.font = FONT_IGNORED
    estado.alignment = CENTER_WRAP
    ws.row_dimensions[i].height = 55

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 32
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 58
ws.freeze_panes = "A2"

out = "Incidencias_Simulador_Respuestas.xlsx"
wb.save(out)
print("Saved:", out, " rows:", len(rows))
