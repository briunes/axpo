"""
Updates 'Incidencias Simulador Signed_reply.xlsx' with statuses and
Spanish developer responses extracted from simulator-issues.html.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# ── Data from simulator-issues.html ──────────────────────────────────────────
# key = issue id (matches column A "Número"), value = (status, respuesta_axpo)

STATUS_ES = {
    "closed":  "Resuelto ✅",
    "open":    "En curso 🔄",
    "ignored": "No procede ❌",
    "ignore":  "No procede ❌",
}

issues = {
    1:  ("ignored", "Las credenciales fueron enviadas correctamente. El acceso fue verificado y funciona. Se descarta la incidencia."),
    2:  ("ignored", "Las credenciales fueron enviadas correctamente. El acceso fue verificado y funciona. Se descarta la incidencia."),
    3:  ("closed",  "Corregido. El importe de ahorro ahora coincide correctamente con la tarifa seleccionada."),
    4:  ("closed",  "Corregido. Los términos de los conceptos de factura han sido actualizados a la terminología correcta: 'Potencia Contraída', 'Potencia Excedente', 'Alquiler de Equipo' e 'Historial de Precios'."),
    5:  ("open",    "En revisión. Se está analizando la coherencia de los precios de potencia de la competencia y la diferencia en el Coste Anual de Potencia."),
    6:  ("closed",  "El consumo anual y la tarifa de acceso pueden obtenerse y completarse manualmente cuando el OCR no los extrae. Resuelto."),
    7:  ("ignored", "Diseño revisado. La pantalla de simulaciones tiene su propia identidad visual y no requiere cambios adicionales."),
    8:  ("ignored", "Funcionalidad de comisiones aplazada. Riesgo identificado: usuarios con el mismo perfil podrían ver datos de comisión de otros — responsabilidad del agente."),
    9:  ("ignored", "Cada agencia ve únicamente sus propios clientes. El administrador puede ver clientes de todas las agencias. El agente ve todos los clientes de su agencia. Los comerciales ven solo los suyos. La importación de leads puede valorarse en una fase futura."),
    10: ("ignored", "Sugerencia registrada. Los precios en €/kWh no se mostrarán en esta versión. Puede considerarse en futuras iteraciones."),
    11: ("closed",  "Resuelto. Los resultados de la comparativa se ordenan ahora de mayor a menor porcentaje de ahorro."),
    12: ("ignored", "La visualización de comisiones en los resultados de simulación queda aplazada para la siguiente fase del proyecto."),
    13: ("closed",  "Corregido. Al editar el nombre del cliente, todos los campos relacionados se actualizan automáticamente."),
    14: ("closed",  "Corregido. Los campos de fecha en la sección de lectura de factura pueden borrarse completamente de una sola vez."),
    15: ("closed",  "Corregido. El campo de importe de factura gestiona correctamente el separador decimal con coma (ej. 54,45 se registra como 54,45)."),
    16: ("closed",  "Corregido. El nombre del agente se autocompleta automáticamente con el usuario que ha iniciado sesión."),
    17: ("closed",  "Resuelto. La generación y descarga del PDF funciona correctamente."),
    18: ("closed",  "Resuelto. Se han añadido los productos 'Gas Estable N2' y 'Gas Estable N3'. La denominación 'Gas Fijo' ha sido sustituida por los nombres de campaña correctos."),
    19: ("closed",  "Resuelto. El OCR posiciona ahora los datos extraídos en los campos correctos."),
    20: ("closed",  "Resuelto. Se ha implementado la subida de facturas mediante arrastrar y soltar (drag & drop)."),
    21: ("closed",  "Resuelto. El CUPS se extrae correctamente de las facturas de Naturgy electricidad."),
    22: ("closed",  "Resuelto. Todos los campos relacionados (tarifa, perfil de carga, periodo de facturación, potencia contratada, consumo energético, totales de factura e impuestos) están ahora agrupados en una única sección colapsable 'Desglose de Factura'."),
    23: ("closed",  "Resuelto. El Perfil de Carga aparece ahora una única vez, dentro de la subsección de Datos de Factura. Se ha añadido una descripción de lo que implica su selección."),
    24: ("closed",  "Resuelto. Las opciones personalizadas (Margen índice, Margen OMIE+B, Margen de potencia) se muestran directamente en la pestaña de Resultados, asociadas a cada fila de resultado correspondiente."),
    25: ("closed",  "Resuelto. Se ha añadido la opción 'Fijo Personalizado' que permite introducir precios de energía y potencia libremente."),
    26: ("closed",  "Resuelto. El campo OMIE €/MWh ha sido retirado de esa sección. Las opciones personalizadas están ahora en la sección de Resultados, ligadas a cada línea de resultado."),
    27: ("closed",  "Resuelto. El historial de precios muestra correctamente los valores mensuales sin repeticiones al seleccionar la tarifa Dinámica Control."),
    28: ("closed",  "Resuelto. La ubicación del botón de descarga del PDF está más clara. El flujo de compartir/descargar ha sido simplificado."),
    29: ("ignored", "Las plantillas PDF personalizadas por agencia, incluyendo la subida de logo, se valorarán en una fase futura del proyecto."),
    30: ("ignored", "El diseño de la simulación final no está basado en el Excel de referencia. Su aspecto actual es intencional y está adaptado al entorno web."),
    31: ("ignored", "La integración con Nemon para obtener el consumo anual automáticamente al introducir el CUPS queda pendiente para una fase futura."),
    32: ("closed",  "Resuelto. El CUPS se autocompleta correctamente, incluso en facturas de Iberdrola donde los dígitos vienen separados."),
    33: ("closed",  "Resuelto. Los excesos de potencia se rellenan automáticamente desde la factura cuando están claramente reflejados."),
    34: ("ignored", "Las dos primeras filas de la comparativa PDF (Potencia Contratada / Consumo Factura) muestran los mismos datos intencionadamente, ya que el cliente no cambia estos valores. Se valora rediseñar esta sección en una versión futura."),
    35: ("closed",  "Resuelto. El campo de potencia permite ahora borrar el valor por defecto (0) y escribir directamente la potencia correcta."),
    36: ("closed",  "Resuelto. El IVA y el Impuesto Eléctrico se presentan ahora como selectores desplegables con opciones configurables desde los Ajustes del Sistema. Si solo existe una opción disponible, se muestra un campo numérico directo."),
    37: ("closed",  "Resuelto. El Perfil de Carga duplicado ha sido eliminado. Ahora aparece únicamente en la sección de Datos de Factura."),
    38: ("closed",  "Resuelto en el motor de cálculo. Los productos 1P_PLUS y 1P_PLUS_XL son de 'Periodo Único' — el precio P1 se aplica a todos los periodos de energía y potencia, funcionando correctamente con la tarifa de acceso 6.1TD (TLV) con una sola entrada de precio por nivel."),
    39: ("closed",  "Resuelto. La última columna de la lista de simulaciones ha sido clarificada. Se ha añadido una opción de descarga directa del PDF desde esa vista."),
    40: ("ignored", "El comportamiento es el previsto: una vez que la simulación ha sido compartida con el cliente, no puede modificarse para garantizar la integridad de la oferta enviada."),
    41: ("closed",  "Resuelto. El PDF incluye ahora el aviso: 'Los precios indicados no están garantizados'."),
    42: ("closed",  "Resuelto. Se han añadido gráficos visuales del ahorro anual y por ciclo de facturación en la vista de resultados."),
    43: ("closed",  "Resuelto. El error de lectura OCR ha sido corregido. El sistema lee correctamente las facturas en múltiples formatos."),
    44: ("closed",  "Resuelto. El PDF exportado incluye ahora el porcentaje de ahorro anual y el importe monetizado del ahorro anual."),
    45: ("ignored", "Confirmado. Los usuarios de agencia ven únicamente sus propias simulaciones. Los agentes ven las simulaciones de su agencia. Los comerciales ven solo las que ellos han creado."),
    46: ("open",    "En revisión. Se está analizando el cálculo incorrecto del importe total de potencia para el caso 'Panificadora Amaya'."),
    47: ("closed",  "Resuelto. El simulador detecta correctamente el mes de facturación desde la factura."),
    48: ("ignored", "Aviso registrado. Los precios de campaña se monitorizan y actualizan periódicamente. Es responsabilidad del equipo mantenerlos al día."),
    49: ("closed",  "Resuelto. Los datos del historial de precios han sido revisados y corregidos."),
    50: ("closed",  "Resuelto. Los periodos de facturación pueden editarse manualmente."),
    51: ("closed",  "Resuelto. La subida de facturas admite ahora múltiples archivos en distintos formatos (no solo un JPEG)."),
}

# ── Colour fills ─────────────────────────────────────────────────────────────
FILL_RESOLVED = PatternFill("solid", fgColor="C6EFCE")   # green
FILL_OPEN     = PatternFill("solid", fgColor="FFEB9C")   # yellow
FILL_IGNORED  = PatternFill("solid", fgColor="E0E0E0")   # grey
FILL_HEADER   = PatternFill("solid", fgColor="1F3864")   # dark blue

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
FONT_RESOLVED = Font(color="276221", bold=True)
FONT_OPEN     = Font(color="9C6500", bold=True)
FONT_IGNORED  = Font(color="595959", bold=True)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")

# ── Load workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook("Incidencias Simulador Signed_reply.xlsx")
ws = wb.active

# ── Add / ensure "Respuesta Axpo" header in column G ─────────────────────────
# Current columns: A=Número B=Incidencia C=Descripción D=Comentarios E=Abierta por F=Estado
header_g = ws.cell(row=1, column=7)
header_g.value = "Respuesta Axpo"
header_g.fill = FILL_HEADER
header_g.font = FONT_HEADER
header_g.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_g.border = BORDER

# Style the existing header row while we're here
for col in range(1, 7):
    cell = ws.cell(row=1, column=col)
    cell.fill = FILL_HEADER
    cell.font = FONT_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

# ── Update each data row ──────────────────────────────────────────────────────
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    num_cell   = row[0]  # A – Número
    estado_cell= row[5]  # F – Estado
    reply_cell = ws.cell(row=num_cell.row, column=7)  # G – Respuesta Axpo

    issue_id = num_cell.value
    if issue_id not in issues:
        continue

    status_key, respuesta = issues[issue_id]
    status_label = STATUS_ES[status_key]

    # Estado
    estado_cell.value = status_label
    if status_key == "closed":
        estado_cell.fill = FILL_RESOLVED
        estado_cell.font = FONT_RESOLVED
    elif status_key == "open":
        estado_cell.fill = FILL_OPEN
        estado_cell.font = FONT_OPEN
    else:
        estado_cell.fill = FILL_IGNORED
        estado_cell.font = FONT_IGNORED
    estado_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    estado_cell.border = BORDER

    # Respuesta Axpo
    reply_cell.value = respuesta
    reply_cell.alignment = WRAP
    reply_cell.border = BORDER

    # Style all other cells in the row for consistency
    for cell in row:
        cell.alignment = WRAP
        cell.border = BORDER

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 38
ws.column_dimensions["D"].width = 38
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 58

# Freeze header row
ws.freeze_panes = "A2"

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = "Incidencias Simulador Signed_reply.xlsx"
wb.save(out_path)
print(f"✅ Guardado: {out_path}")
print(f"   Filas actualizadas: {len(issues)}")
